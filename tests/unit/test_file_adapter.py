"""
tests/unit/test_file_adapter.py
Unit tests for nexus/source/file/adapter.py.

All tests inject mock readers via the pluggable factory parameters so
that no real browser, file system I/O, or external library binaries are
required in the unit layer.

A separate "integration-style" section near the bottom uses the
conftest.py file fixtures (real pypdf / openpyxl) to validate that the
default factories work end-to-end.
"""
from __future__ import annotations

from pathlib import Path
from typing import Any
from unittest.mock import MagicMock

import pytest

from nexus.source.file.adapter import DocumentContent, FileAdapter

# ---------------------------------------------------------------------------
# Mock factory helpers
# ---------------------------------------------------------------------------


def _mock_pdf_page(text: str) -> MagicMock:
    page = MagicMock()
    page.extract_text.return_value = text
    return page


def _pdf_factory(*, is_encrypted: bool = False, pages_text: list[str] | None = None) -> Any:
    """Return a factory that yields a mock pypdf-style reader."""
    reader = MagicMock()
    reader.is_encrypted = is_encrypted
    reader.pages = [_mock_pdf_page(t) for t in (pages_text or [])]
    return lambda path: reader


def _failing_pdf_factory(exc: Exception | None = None) -> Any:
    """Return a factory that raises when called."""
    def _factory(path: Path) -> Any:
        raise exc or OSError("file not found")
    return _factory


def _ocr_factory(pages: list[str]) -> Any:
    return lambda path: pages


def _failing_ocr_factory(exc: Exception | None = None) -> Any:
    def _factory(path: Path) -> list[str]:
        raise exc or RuntimeError("OCR failed")
    return _factory


class _MockSheet:
    def __init__(self, title: str, rows: list[tuple]) -> None:
        self.title = title
        self._rows = rows

    def iter_rows(self, *, values_only: bool = False):
        return iter(self._rows)


class _MockWorkbook:
    def __init__(self, sheets: list[_MockSheet]) -> None:
        self.worksheets = sheets


def _xlsx_factory(sheets: list[_MockSheet]) -> Any:
    wb = _MockWorkbook(sheets)
    return lambda path: wb


def _failing_xlsx_factory(exc: Exception | None = None) -> Any:
    def _factory(path: Path) -> Any:
        raise exc or OSError("file not found")
    return _factory


# ---------------------------------------------------------------------------
# is_supported
# ---------------------------------------------------------------------------


class TestIsSupported:
    def test_pdf_is_supported(self):
        adapter = FileAdapter()
        assert adapter.is_supported("report.pdf") is True

    def test_xlsx_is_supported(self):
        adapter = FileAdapter()
        assert adapter.is_supported("data.xlsx") is True

    def test_xls_is_supported(self):
        adapter = FileAdapter()
        assert adapter.is_supported("old.xls") is True

    def test_txt_is_not_supported(self):
        adapter = FileAdapter()
        assert adapter.is_supported("notes.txt") is False

    def test_docx_is_not_supported(self):
        adapter = FileAdapter()
        assert adapter.is_supported("doc.docx") is False

    def test_case_insensitive(self):
        adapter = FileAdapter()
        assert adapter.is_supported("report.PDF") is True
        assert adapter.is_supported("DATA.XLSX") is True


# ---------------------------------------------------------------------------
# extract() — unsupported type
# ---------------------------------------------------------------------------


class TestExtractUnsupported:
    def test_returns_none_for_txt(self):
        adapter = FileAdapter()
        assert adapter.extract("notes.txt") is None

    def test_returns_none_for_docx(self):
        adapter = FileAdapter()
        assert adapter.extract("doc.docx") is None

    def test_returns_none_for_png(self):
        adapter = FileAdapter()
        assert adapter.extract("image.png") is None


# ---------------------------------------------------------------------------
# extract() — PDF text path
# ---------------------------------------------------------------------------


class TestExtractPdfText:
    def test_text_pdf_returns_pdf_text_source_type(self):
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(pages_text=["Hello World"])
        )
        result = adapter.extract("report.pdf")
        assert result is not None
        assert result.source_type == "pdf_text"

    def test_text_pdf_returns_correct_pages(self):
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(pages_text=["Page one", "Page two"])
        )
        result = adapter.extract("report.pdf")
        assert result is not None
        assert result.pages == ["Page one", "Page two"]

    def test_text_pdf_confidence_is_1(self):
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(pages_text=["text"])
        )
        result = adapter.extract("report.pdf")
        assert result is not None
        assert result.extraction_confidence == 1.0

    def test_text_pdf_tables_is_empty(self):
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(pages_text=["text"])
        )
        result = adapter.extract("report.pdf")
        assert result is not None
        assert result.tables == []

    def test_text_pdf_metadata_has_page_count(self):
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(pages_text=["a", "b"])
        )
        result = adapter.extract("report.pdf")
        assert result is not None
        assert result.metadata["page_count"] == 2

    def test_whitespace_only_pages_still_route_to_text(self):
        # strip() is applied — blank strings count as no-text per page
        # but if at least one stripped page is non-empty → pdf_text
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(pages_text=["  ", "Real content"])
        )
        result = adapter.extract("doc.pdf")
        assert result is not None
        assert result.source_type == "pdf_text"


# ---------------------------------------------------------------------------
# extract() — PDF encrypted path
# ---------------------------------------------------------------------------


class TestExtractPdfEncrypted:
    def test_encrypted_pdf_returns_none(self):
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(is_encrypted=True)
        )
        result = adapter.extract("secret.pdf")
        assert result is None

    def test_encrypted_pdf_does_not_call_ocr(self):
        ocr_called = []
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(is_encrypted=True),
            _ocr_reader_factory=lambda p: ocr_called.append(p) or [],
        )
        adapter.extract("secret.pdf")
        assert ocr_called == []


# ---------------------------------------------------------------------------
# extract() — PDF OCR fallback path
# ---------------------------------------------------------------------------


class TestExtractPdfOcr:
    def test_empty_text_falls_back_to_ocr(self):
        """All pages empty → OCR path → pdf_ocr."""
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(pages_text=["", ""]),
            _ocr_reader_factory=_ocr_factory(["OCR text"]),
        )
        result = adapter.extract("scan.pdf")
        assert result is not None
        assert result.source_type == "pdf_ocr"

    def test_ocr_pages_are_returned(self):
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(pages_text=[""]),
            _ocr_reader_factory=_ocr_factory(["Recognised text here"]),
        )
        result = adapter.extract("scan.pdf")
        assert result is not None
        assert result.pages == ["Recognised text here"]

    def test_ocr_confidence_is_0_85(self):
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(pages_text=[""]),
            _ocr_reader_factory=_ocr_factory(["text"]),
        )
        result = adapter.extract("scan.pdf")
        assert result is not None
        assert result.extraction_confidence == pytest.approx(0.85)

    def test_ocr_failure_returns_none(self):
        adapter = FileAdapter(
            _pdf_reader_factory=_pdf_factory(pages_text=[""]),
            _ocr_reader_factory=_failing_ocr_factory(),
        )
        assert adapter.extract("scan.pdf") is None


# ---------------------------------------------------------------------------
# extract() — PDF error path
# ---------------------------------------------------------------------------


class TestExtractPdfErrors:
    def test_open_failure_returns_none(self):
        adapter = FileAdapter(_pdf_reader_factory=_failing_pdf_factory())
        assert adapter.extract("missing.pdf") is None

    def test_page_extract_exception_uses_empty_string(self):
        """If one page's extract_text raises, that page becomes empty string."""
        bad_page = MagicMock()
        bad_page.extract_text.side_effect = RuntimeError("corrupt page")

        good_page = MagicMock()
        good_page.extract_text.return_value = "Good content"

        reader = MagicMock()
        reader.is_encrypted = False
        reader.pages = [bad_page, good_page]

        adapter = FileAdapter(_pdf_reader_factory=lambda p: reader)
        result = adapter.extract("doc.pdf")

        assert result is not None
        assert result.source_type == "pdf_text"
        assert result.pages[0] == ""
        assert result.pages[1] == "Good content"


# ---------------------------------------------------------------------------
# extract() — XLSX path
# ---------------------------------------------------------------------------


class TestExtractXlsx:
    def _make_adapter(self, sheets: list[_MockSheet]) -> FileAdapter:
        return FileAdapter(_xlsx_reader_factory=_xlsx_factory(sheets))

    def test_xlsx_returns_xlsx_source_type(self):
        sheet = _MockSheet("Sheet1", [("A", "B"), ("1", "2")])
        result = self._make_adapter([sheet]).extract("data.xlsx")
        assert result is not None
        assert result.source_type == "xlsx"

    def test_xlsx_pages_are_sheet_titles(self):
        sheets = [_MockSheet("Sales", []), _MockSheet("Costs", [])]
        result = self._make_adapter(sheets).extract("data.xlsx")
        assert result is not None
        assert result.pages == ["Sales", "Costs"]

    def test_xlsx_tables_contain_rows(self):
        sheet = _MockSheet("Data", [("Product", "Q1"), ("Widget", 100)])
        result = self._make_adapter([sheet]).extract("data.xlsx")
        assert result is not None
        assert result.tables == [[["Product", "Q1"], ["Widget", "100"]]]

    def test_xlsx_none_cells_become_empty_string(self):
        sheet = _MockSheet("Sheet1", [(None, "value", None)])
        result = self._make_adapter([sheet]).extract("data.xlsx")
        assert result is not None
        assert result.tables[0][0] == ["", "value", ""]

    def test_xlsx_confidence_is_1(self):
        sheet = _MockSheet("S", [("x",)])
        result = self._make_adapter([sheet]).extract("data.xlsx")
        assert result is not None
        assert result.extraction_confidence == 1.0

    def test_xlsx_metadata_has_sheet_count(self):
        sheets = [_MockSheet("A", []), _MockSheet("B", [])]
        result = self._make_adapter(sheets).extract("data.xlsx")
        assert result is not None
        assert result.metadata["sheet_count"] == 2

    def test_xlsx_open_failure_returns_none(self):
        adapter = FileAdapter(_xlsx_reader_factory=_failing_xlsx_factory())
        assert adapter.extract("missing.xlsx") is None

    def test_xls_extension_is_handled(self):
        sheet = _MockSheet("Sheet1", [("val",)])
        result = self._make_adapter([sheet]).extract("legacy.xls")
        assert result is not None
        assert result.source_type == "xlsx"

    def test_multiple_sheets_produce_multiple_tables(self):
        sheets = [
            _MockSheet("S1", [("a",), ("b",)]),
            _MockSheet("S2", [("c",)]),
        ]
        result = self._make_adapter(sheets).extract("multi.xlsx")
        assert result is not None
        assert len(result.tables) == 2
        assert result.tables[0] == [["a"], ["b"]]
        assert result.tables[1] == [["c"]]


# ---------------------------------------------------------------------------
# DocumentContent data model
# ---------------------------------------------------------------------------


class TestDocumentContent:
    def test_can_construct_directly(self):
        dc = DocumentContent(
            source_type="pdf_text",
            pages=["page 1"],
            tables=[],
            metadata={},
            extraction_confidence=1.0,
        )
        assert dc.source_type == "pdf_text"
        assert dc.pages == ["page 1"]

    def test_all_source_types_are_valid(self):
        for stype in ("pdf_text", "pdf_ocr", "xlsx", "direct"):
            dc = DocumentContent(
                source_type=stype,  # type: ignore[arg-type]
                pages=[],
                tables=[],
                metadata={},
                extraction_confidence=0.9,
            )
            assert dc.source_type == stype


# ---------------------------------------------------------------------------
# Integration-style tests using real fixture files (conftest.py)
# ---------------------------------------------------------------------------


class TestRealFixtures:
    """
    These tests exercise the *real* pypdf and openpyxl reader stacks.
    Fixtures are provided by conftest.py and written to tmp_path.
    """

    def test_text_pdf_fixture_routes_to_pdf_text(self, test_pdf: Path):
        """Real pypdf reader: text PDF → pdf_text."""
        import pypdf

        def real_pdf_reader(path: Path) -> Any:
            return pypdf.PdfReader(str(path))

        adapter = FileAdapter(_pdf_reader_factory=real_pdf_reader)
        result = adapter.extract(test_pdf)

        assert result is not None
        assert result.source_type == "pdf_text"
        assert any("Hello World" in p for p in result.pages)

    def test_encrypted_pdf_fixture_returns_none(self, test_scanned_pdf: Path):
        """Real pypdf reader: encrypted PDF → None (HITL signal emitted)."""
        import pypdf

        def real_pdf_reader(path: Path) -> Any:
            return pypdf.PdfReader(str(path))

        adapter = FileAdapter(_pdf_reader_factory=real_pdf_reader)
        result = adapter.extract(test_scanned_pdf)

        assert result is None

    def test_xlsx_fixture_table_content(self, test_xlsx: Path):
        """Real openpyxl reader: correct table data from multi-sheet XLSX."""
        import openpyxl

        def real_xlsx_reader(path: Path) -> Any:
            return openpyxl.load_workbook(str(path), data_only=True)

        adapter = FileAdapter(_xlsx_reader_factory=real_xlsx_reader)
        result = adapter.extract(test_xlsx)

        assert result is not None
        assert result.source_type == "xlsx"
        assert result.pages == ["Sales", "Costs"]

        # Sales sheet: header + 2 data rows
        sales = result.tables[0]
        assert sales[0] == ["Product", "Q1", "Q2"]
        assert sales[1] == ["Widget", "100", "200"]
        assert sales[2] == ["Gadget", "150", "175"]

        # Costs sheet: header + 1 data row
        costs = result.tables[1]
        assert costs[0] == ["Item", "Amount"]
        assert costs[1] == ["Rent", "5000"]
