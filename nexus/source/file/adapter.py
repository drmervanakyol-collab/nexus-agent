"""
nexus/source/file/adapter.py
File / Document adapter for Nexus Agent.

Architecture
------------
FileAdapter extracts content from PDF and XLSX files into a
normalised DocumentContent structure.  A pluggable reader-factory
pattern (identical to the DOM adapter's ``_session_factory``) lets
unit tests inject mock objects without touching real files or system
libraries.

PDF extraction path
-------------------
1. Open with pypdf.PdfReader via ``_pdf_reader_factory``.
2. If the file is encrypted → return None, emit a structured HITL
   warning via the logger (``hitl_required=True``).
3. Extract text from every page.  If *any* page has non-empty text
   → ``source_type = "pdf_text"``, confidence 1.0.
4. No text anywhere → call ``_ocr_reader_factory`` → ``source_type
   = "pdf_ocr"``, confidence 0.85.
5. Any unhandled exception → log + return None.

XLSX extraction path
--------------------
Open with openpyxl.load_workbook via ``_xlsx_reader_factory``.
``source_type = "xlsx"``, pages = sheet titles,
tables = cell values per sheet.

Unsupported file types → None.

Default reader factories
------------------------
_default_pdf_reader_factory  — lazy ``import pypdf``
_default_ocr_reader_factory  — lazy ``import pytesseract`` + ``pdf2image``
_default_xlsx_reader_factory — lazy ``import openpyxl``

All lazy imports are deferred inside the factory so that tests that
inject mocks never trigger an ImportError for missing system packages.
"""
from __future__ import annotations

from collections.abc import Callable
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Literal

from nexus.infra.logger import get_logger

_log = get_logger(__name__)

# ---------------------------------------------------------------------------
# Public data model
# ---------------------------------------------------------------------------


@dataclass
class DocumentContent:
    """
    Normalised content extracted from a file.

    Attributes
    ----------
    source_type:
        How the content was obtained:
        ``"pdf_text"``  — text layer from a native-text PDF
        ``"pdf_ocr"``   — OCR run on a scanned / image-only PDF
        ``"xlsx"``      — spreadsheet extracted via openpyxl
        ``"direct"``    — raw text passed in directly (reserved for future use)
    pages:
        Per-page text strings (PDF) or sheet titles (XLSX).
    tables:
        List of tables; each table is a list of rows, each row a list
        of cell strings.  Empty for ``"pdf_text"`` and ``"pdf_ocr"``.
    metadata:
        Arbitrary key-value pairs (``page_count``, ``path``, …).
    extraction_confidence:
        Float in [0, 1].  1.0 for deterministic extraction, 0.85 for OCR.
    """

    source_type: Literal["pdf_text", "pdf_ocr", "xlsx", "direct"]
    pages: list[str]
    tables: list[list[list[str]]]
    metadata: dict[str, Any]
    extraction_confidence: float


# ---------------------------------------------------------------------------
# Type aliases for pluggable factories
# ---------------------------------------------------------------------------

# callable(path: Path) -> Any  (a pypdf.PdfReader-like object)
_PdfReaderFactory = Callable[[Path], Any]

# callable(path: Path) -> list[str]  (OCR'd text per page)
_OcrReaderFactory = Callable[[Path], list[str]]

# callable(path: Path) -> Any  (an openpyxl Workbook-like object)
_XlsxReaderFactory = Callable[[Path], Any]

# ---------------------------------------------------------------------------
# Default reader factories — lazy imports so mocks never need these deps
# ---------------------------------------------------------------------------

_SUPPORTED_SUFFIXES: frozenset[str] = frozenset({".pdf", ".xlsx", ".xls"})


def _default_pdf_reader_factory(path: Path) -> Any:
    import pypdf  # noqa: PLC0415

    return pypdf.PdfReader(str(path))


def _default_ocr_reader_factory(path: Path) -> list[str]:
    """
    Convert PDF pages to images and run Tesseract OCR on each.

    Requires ``pytesseract``, ``Pillow``, and ``pdf2image`` (plus the
    Tesseract binary).  Raises if any dependency is absent.
    """
    import pytesseract  # noqa: PLC0415, I001
    from pdf2image import convert_from_path  # noqa: PLC0415

    images = convert_from_path(str(path))
    return [pytesseract.image_to_string(img) for img in images]


def _default_xlsx_reader_factory(path: Path) -> Any:
    import openpyxl  # noqa: PLC0415  # type: ignore[import-untyped]

    return openpyxl.load_workbook(str(path), data_only=True)


# ---------------------------------------------------------------------------
# FileAdapter
# ---------------------------------------------------------------------------


class FileAdapter:
    """
    Synchronous adapter for extracting structured content from files.

    Parameters
    ----------
    _pdf_reader_factory:
        Callable ``(path) -> reader``.  The reader must expose
        ``.is_encrypted: bool`` and ``.pages: Iterable`` where each
        page has ``.extract_text() -> str``.
        Defaults to the real pypdf stack.
    _ocr_reader_factory:
        Callable ``(path) -> list[str]`` — one string per page.
        Defaults to the pytesseract stack.
    _xlsx_reader_factory:
        Callable ``(path) -> workbook``.  The workbook must expose
        ``.worksheets: list`` where each sheet has ``.title: str``
        and ``.iter_rows(values_only=True) -> Iterable[tuple]``.
        Defaults to the openpyxl stack.
    """

    def __init__(
        self,
        *,
        _pdf_reader_factory: _PdfReaderFactory | None = None,
        _ocr_reader_factory: _OcrReaderFactory | None = None,
        _xlsx_reader_factory: _XlsxReaderFactory | None = None,
    ) -> None:
        self._pdf_reader = _pdf_reader_factory or _default_pdf_reader_factory
        self._ocr_reader = _ocr_reader_factory or _default_ocr_reader_factory
        self._xlsx_reader = _xlsx_reader_factory or _default_xlsx_reader_factory

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def is_supported(self, file_path: str | Path) -> bool:
        """Return True if the file extension is supported."""
        return Path(file_path).suffix.lower() in _SUPPORTED_SUFFIXES

    def extract(self, file_path: str | Path) -> DocumentContent | None:
        """
        Extract content from *file_path*.

        Returns
        -------
        DocumentContent on success, None on any failure (unsupported
        type, encrypted PDF, or unhandled exception).
        """
        path = Path(file_path)
        suffix = path.suffix.lower()

        if suffix == ".pdf":
            return self._extract_pdf(path)
        if suffix in {".xlsx", ".xls"}:
            return self._extract_xlsx(path)

        _log.debug("file_adapter_unsupported_type", suffix=suffix, path=str(path))
        return None

    # ------------------------------------------------------------------
    # PDF path
    # ------------------------------------------------------------------

    def _extract_pdf(self, path: Path) -> DocumentContent | None:
        try:
            reader = self._pdf_reader(path)
        except Exception as exc:
            _log.debug("pdf_open_failed", path=str(path), error=str(exc))
            return None

        # Encrypted gate — emit HITL signal and abort
        if reader.is_encrypted:
            _log.warning(
                "pdf_encrypted_hitl_required",
                path=str(path),
                hitl_required=True,
                reason="encrypted_pdf",
            )
            return None

        # Extract text layer
        pages: list[str] = []
        for page in reader.pages:
            try:
                raw = page.extract_text() or ""
                pages.append(raw.strip())
            except Exception as exc:
                _log.debug(
                    "pdf_page_extract_failed", path=str(path), error=str(exc)
                )
                pages.append("")

        if any(pages):  # at least one page has text
            return DocumentContent(
                source_type="pdf_text",
                pages=pages,
                tables=[],
                metadata={"page_count": len(pages), "path": str(path)},
                extraction_confidence=1.0,
            )

        # No text layer → OCR fallback
        try:
            ocr_pages = self._ocr_reader(path)
            return DocumentContent(
                source_type="pdf_ocr",
                pages=ocr_pages,
                tables=[],
                metadata={"page_count": len(ocr_pages), "path": str(path)},
                extraction_confidence=0.85,
            )
        except Exception as exc:
            _log.debug("pdf_ocr_failed", path=str(path), error=str(exc))
            return None

    # ------------------------------------------------------------------
    # XLSX path
    # ------------------------------------------------------------------

    def _extract_xlsx(self, path: Path) -> DocumentContent | None:
        try:
            workbook = self._xlsx_reader(path)
        except Exception as exc:
            _log.debug("xlsx_open_failed", path=str(path), error=str(exc))
            return None

        pages: list[str] = []
        tables: list[list[list[str]]] = []

        for sheet in workbook.worksheets:
            pages.append(sheet.title)
            rows: list[list[str]] = []
            for row in sheet.iter_rows(values_only=True):
                rows.append(
                    [str(cell) if cell is not None else "" for cell in row]
                )
            tables.append(rows)

        return DocumentContent(
            source_type="xlsx",
            pages=pages,
            tables=tables,
            metadata={
                "sheet_count": len(workbook.worksheets),
                "path": str(path),
            },
            extraction_confidence=1.0,
        )
